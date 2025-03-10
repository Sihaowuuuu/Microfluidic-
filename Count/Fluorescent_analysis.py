import os
import shutil

import cv2
import numpy as np

from Count import count_channel


def analysis_intensity_over_time(root_path):
    # Dictionary to temporarily store data
    data_dict = {}

    # Iterate over time folders
    for time_folder in sorted(os.listdir(root_path)):  # Sort by time
        if time_folder.startswith("t"):  # Only process folders starting with "t"
            time = time_folder  # Extract time point (e.g., t00 -> 0)
            time_path = os.path.join(root_path, time_folder)

            # Iterate over channel folders
            for channel_folder in sorted(os.listdir(time_path)):
                if channel_folder.startswith("Thr"):
                    print(f"Processing folder: {channel_folder}")

                    channel_parts = channel_folder.split("_")

                    # 通过遍历找到 chXX 这样的字符串
                    channel = None
                    for part in channel_parts:
                        if (
                            part.startswith("ch") and part[2:].isdigit()
                        ):  # 确保 "ch" 后跟的是数字
                            channel = part
                            break

                    if channel not in ["ch01", "ch02"]:
                        continue  # Skip this iteration for unrecognized channels

                    channel_path = os.path.join(time_path, channel_folder)

                    # Iterate over image files
                    for img_file in sorted(os.listdir(channel_path)):
                        img_path = os.path.join(channel_path, img_file)

                        # Ensure it's a file and not a subfolder
                        if os.path.isfile(img_path):
                            tube = img_file.split(".")[0]  # Extract tube info

                            # Initialize keys for the dictionary
                            key = (time, channel, tube)
                            if key not in data_dict:
                                data_dict[key] = {"Intensity": None, "Cell Count": None}

                            # Read the image and process based on folder type
                            img = cv2.imread(img_path, cv2.IMREAD_GRAYSCALE)
                            if img is not None:
                                if channel_folder.startswith("Thr"):
                                    # Calculate fluorescence intensity (mean pixel value)
                                    data_dict[key]["Intensity"] = np.mean(img)
                                    data_dict[key]["Cell Count"] = count_channel(
                                        img_path
                                    )

    # Convert dictionary to a DataFrame
    rows = [
        {
            "Time": time,
            "Channel": channel,
            "Tube": tube,
            "Intensity": values["Intensity"],
            "Cell Count": values["Cell Count"],
        }
        for (time, channel, tube), values in data_dict.items()
    ]
    df = pd.DataFrame(rows)
    return df


def save_to_excel(df, root_path):
    tubes = sorted(df["Tube"].unique())
    times = sorted(df["Time"].unique())
    channels = sorted(df["Channel"].unique())

    header = ["Tube"] + [
        f"{time}"
        for time in times
        for _ in channels
        for _ in ["Cell Count", "Intensity"]
    ]
    sub_header = [""] + [
        f"{ch}_{metric}"
        for time in times
        for ch in channels
        for metric in ["Cell Count", "Intensity"]
    ]

    excel_data = [header, sub_header]
    for tube in tubes:
        row = [tube]
        for time in times:
            for channel in channels:
                group_data = df[
                    (df["Tube"] == tube)
                    & (df["Time"] == time)
                    & (df["Channel"] == channel)
                ]
                # print(f"Tube: {tube}, Time: {time}, Channel: {channel}, Data: {group_data}")  # Debug
                if not group_data.empty:
                    cell_count = group_data.iloc[0].get("Cell Count", 0)
                    intensity = group_data.iloc[0].get("Intensity", 0)
                else:
                    cell_count = 0
                    intensity = 0
                row.extend([cell_count, intensity])
        excel_data.append(row)

    excel_df = pd.DataFrame(excel_data)
    excel_path = os.path.join(root_path, "fluorescent_analysis_results.xlsx")
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        excel_df.to_excel(writer, index=False, header=False)

    print(f"Excel file saved to: {excel_path}")


"""
    # Plot fluorescence intensity over time
    for (tube, channel), group_data in df.groupby(["Tube", "Channel"]):
        plt.figure()
        group_data = group_data.sort_values("Time")  # Sort by time
        plt.plot(group_data["Time"], group_data["Intensity"], marker="o", label="Intensity")
        plt.plot(group_data["Time"], group_data["Cell Count"], marker="x", label="Cell Count")
        plt.title(f"Tube {tube}, Channel {channel}")
        plt.xlabel("Time")
        plt.ylabel("Value")
        plt.legend()
        plt.grid(True)
        # Save the plot to root_path
        plot_path = os.path.join(root_path, f"Tube_{tube}_Channel_{channel}.png")
        plt.savefig(plot_path)
        plt.show()
        print(f"Plot saved to {plot_path}")
"""


def organize_files_by_time_part(source_folder):
    """
    Organizes files into subfolders based on the fifth part of their file names (split by underscore).

    Args:
        source_folder (str): The path to the folder containing the files to be organized.

    """
    # Ensure the source folder exists
    if not os.path.exists(source_folder):
        print(f"Error: The folder '{source_folder}' does not exist.")
        return

    # Iterate over all files in the source folder
    for file_name in os.listdir(source_folder):
        file_path = os.path.join(source_folder, file_name)

        # Skip directories, process only files
        if not os.path.isfile(file_path):
            continue

        # Split the file name by underscore
        parts = file_name.split("_")

        # Identify the time part (e.g., 't00', 't01')
        time_part = None
        for part in parts:
            if (
                part.startswith("t") and part[1:].isdigit()
            ):  # Check if it matches the "t" + digits pattern
                time_part = part
                break

        if not time_part:
            print(f"Skipping '{file_name}': no valid time part found.")
            continue

        # Use the time part as the new folder name
        new_folder_name = time_part
        new_folder_path = os.path.join(source_folder, new_folder_name)

        # Create the folder if it doesn't exist
        if not os.path.exists(new_folder_path):
            os.makedirs(new_folder_path)

        # Move the file into the corresponding folder
        shutil.move(file_path, os.path.join(new_folder_path, file_name))
        print(f"Moved '{file_name}' to '{new_folder_path}'")

    print("File organization complete.")


def group_tubes_by_specific_ratio(df):
    """
    based on the ratio of ch01 and ch02 group

    Args:
        df (pd.DataFrame): contain Time、Channel、Tube、Cell Count's DataFrame。

    Returns:
        dict: 比值作为键，每个比值对应的 Tube 列表。
    """
    ratio_groups = defaultdict(list)  # save the Tube list

    # travel throughout tube list
    for tube in df["Tube"].unique():
        # get Cell Count of tube in ch01 and ch02
        tube_data = df[
            (df["Time"] == "t00")
            & (df["Tube"] == tube)
            & (df["Channel"].isin(["ch01", "ch02"]))
        ]
        if len(tube_data) == 2:  # make sure two channels
            count_ch01 = tube_data[tube_data["Channel"] == "ch01"]["Cell Count"].values[
                0
            ]
            count_ch02 = tube_data[tube_data["Channel"] == "ch02"]["Cell Count"].values[
                0
            ]

            if count_ch02 > 0 and count_ch01 > 0:  # make sure dominator
                # Calculate the ratio and reduce to the simplest fraction
                count_ch01 = int(count_ch01)
                count_ch02 = int(count_ch02)

                ratio = Fraction(count_ch01, count_ch02)
                ratio_groups[(ratio.numerator, ratio.denominator)].append(tube)

    return ratio_groups


def plot_cell_count_over_time_by_ratio(df, root_path):
    """
    plot the line graph based on the ratio.
    """
    # use group_tubes_by_specific_ratio function to group
    ratio_groups = group_tubes_by_specific_ratio(df)

    # create a excel file
    wb = Workbook()

    # travel throughout every ratio
    for ratio, tubes in ratio_groups.items():
        sheet_name = f"Ratio_{ratio[0]}_{ratio[1]}"
        sheet = wb.create_sheet(title=sheet_name)
        plt.figure()  # Creating a new graphic
        plt.title(f"Cell Count Over Time for Ratio {ratio[0]}:{ratio[1]}")

        # Iterate through each test tube and plot a composite graph
        for tube in tubes:
            # Get the data for this tube at ch01 and ch02.
            ch01_data = df[(df["Tube"] == tube) & (df["Channel"] == "ch01")]
            ch02_data = df[(df["Tube"] == tube) & (df["Channel"] == "ch02")]

            # Ensure that data are available
            if not ch01_data.empty and not ch02_data.empty:
                # Sorted data
                ch01_data = ch01_data.sort_values("Time")
                ch02_data = ch02_data.sort_values("Time")

                # Plotting cell counts over time
                plt.plot(
                    ch01_data["Time"],
                    ch01_data["Cell Count"],
                    marker="o",
                    label=f"Tube {tube} - ch01",
                )
                plt.plot(
                    ch02_data["Time"],
                    ch02_data["Cell Count"],
                    marker="x",
                    label=f"Tube {tube} - ch02",
                )

        plt.xlabel("Time")
        plt.ylabel("Cell Count")
        plt.legend()
        plt.grid(True)

        # Save composite image to memory
        img_stream = BytesIO()
        plt.savefig(img_stream, format="png")
        img_stream.seek(0)  # Reset the position of the stream
        img = Image(img_stream)

        # Insert the composite image into the ratio sheet
        sheet.add_image(img, "A1")

        plt.close()  # Close Integrated Graphics

        # Generate a separate image for each tube and insert it into the corresponding ratio sheet
        current_row = 30  # Insert tube image starting in cell A20
        for tube in tubes:
            tube_sheet_title = f"Tube {tube} - Cell Count"
            plt.figure()
            plt.title(f"Cell Count Over Time for Tube {tube}")

            # Get the tube's data at ch01 and ch02
            ch01_data = df[(df["Tube"] == tube) & (df["Channel"] == "ch01")]
            ch02_data = df[(df["Tube"] == tube) & (df["Channel"] == "ch02")]

            # Plotting cell counts over time
            if not ch01_data.empty and not ch02_data.empty:
                # Plotting cell counts over time
                ch01_data = ch01_data.sort_values("Time")
                ch02_data = ch02_data.sort_values("Time")

                plt.plot(
                    ch01_data["Time"],
                    ch01_data["Cell Count"],
                    marker="o",
                    label=f"ch01",
                )
                plt.plot(
                    ch02_data["Time"],
                    ch02_data["Cell Count"],
                    marker="x",
                    label=f"ch02",
                )

            plt.xlabel("Time")
            plt.ylabel("Cell Count")
            plt.legend()
            plt.grid(True)

            # Save individual tube images to memory
            img_stream_tube = BytesIO()
            plt.savefig(img_stream_tube, format="png")
            img_stream_tube.seek(0)
            img_tube = Image(img_stream_tube)

            # Insert tube image into ratio sheet (starting with A20)
            sheet.add_image(img_tube, f"A{current_row}")
            current_row += 30  # Each tube image is spaced 30 lines apart

            plt.close()  # Close the tube's graphic

    # Save Excel file
    excel_path = os.path.join(root_path, "fluorescent_analysis_results_graph.xlsx")
    excel_dir = os.path.dirname(excel_path)

    # If the directory does not exist, create the directory
    if not os.path.exists(excel_dir):
        os.makedirs(excel_dir)
    wb.save(excel_path)
    print(f"Excel file saved to {excel_path}")


import matplotlib.pyplot as plt
import pandas as pd
from collections import defaultdict
from fractions import Fraction
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from io import BytesIO


def plot_intensity_over_time_by_ratio(df, excel_path):
    """
    Group the tubes according to the ratio and plot the intensity of the different channels at each ratio over time.
    Insert the image for each ratio into a different sheet in Excel, as well as generating a separate image for each Tube and inserting it into the appropriate ratio sheet.
    """
    # Grouping using the previous group_tubes_by_specific_ratio function
    ratio_groups = group_tubes_by_specific_ratio(df)

    # Creating Excel Files
    wb = Workbook()

    # Iterate over each ratio
    for ratio, tubes in ratio_groups.items():
        sheet_name = f"Ratio_{ratio[0]}_{ratio[1]}"
        sheet = wb.create_sheet(title=sheet_name)
        plt.figure()  # Creating a new graphic
        plt.title(f"Intensity Over Time for Ratio {ratio[0]}:{ratio[1]}")

        # Iterate through each test tube and plot a composite graph
        for tube in tubes:
            # Get the data for this tube at ch01 and ch02.
            ch01_data = df[(df["Tube"] == tube) & (df["Channel"] == "ch01")]
            ch02_data = df[(df["Tube"] == tube) & (df["Channel"] == "ch02")]

            # Ensure availability of data
            if not ch01_data.empty and not ch02_data.empty:
                # Sorted data
                ch01_data = ch01_data.sort_values("Time")
                ch02_data = ch02_data.sort_values("Time")

                # Plotting intensity over time
                plt.plot(
                    ch01_data["Time"],
                    ch01_data["Intensity"],
                    marker="o",
                    label=f"Tube {tube} - ch01",
                )
                plt.plot(
                    ch02_data["Time"],
                    ch02_data["Intensity"],
                    marker="x",
                    label=f"Tube {tube} - ch02",
                )

        plt.xlabel("Time")
        plt.ylabel("Intensity")
        plt.legend()
        plt.grid(True)

        # Save composite image to memory
        img_stream = BytesIO()
        plt.savefig(img_stream, format="png")
        img_stream.seek(0)  # Reset the position of the stream
        img = Image(img_stream)

        # Insert the composite image into the ratio sheet
        sheet.add_image(img, "A1")

        plt.close()  # Close Integrated Graphics

        # Generate a separate image for each tube and insert it into the corresponding ratio sheet
        current_row = 20  # Insert tube image starting in cell A20
        for tube in tubes:
            tube_sheet_title = f"Tube {tube} - Intensity"
            plt.figure()  # Creating a new graphic
            plt.title(f"Intensity Over Time for Tube {tube}")

            # Get the tube's data at ch01 and ch02.
            ch01_data = df[(df["Tube"] == tube) & (df["Channel"] == "ch01")]
            ch02_data = df[(df["Tube"] == tube) & (df["Channel"] == "ch02")]

            # Plotting intensity over time
            if not ch01_data.empty and not ch02_data.empty:
                # Sorted data
                ch01_data = ch01_data.sort_values("Time")
                ch02_data = ch02_data.sort_values("Time")

                plt.plot(
                    ch01_data["Time"], ch01_data["Intensity"], marker="o", label=f"ch01"
                )
                plt.plot(
                    ch02_data["Time"], ch02_data["Intensity"], marker="x", label=f"ch02"
                )

            plt.xlabel("Time")
            plt.ylabel("Intensity")
            plt.legend()
            plt.grid(True)

            # Save individual tube images to memory
            img_stream_tube = BytesIO()
            plt.savefig(img_stream_tube, format="png")
            img_stream_tube.seek(0)  # Reset the position of the stream
            img_tube = Image(img_stream_tube)

            # Insert tube image into ratio sheet (starting with A20)
            sheet.add_image(img_tube, f"A{current_row}")
            current_row += 20  # Each tube image is spaced 20 lines apart

            plt.close()  # Turning off the tube's graphics
    # Save Excel file
    excel_path = os.path.join(
        excel_path, "fluorescent_analysis_results_intensity_graph.xlsx"
    )
    excel_dir = os.path.dirname(excel_path)

    # If the directory does not exist, create the directory
    if not os.path.exists(excel_dir):
        os.makedirs(excel_dir)
    wb.save(excel_path)
    print(f"Excel file saved to {excel_path}")


def plot_normalized_intensity_over_time_by_ratio(df, excel_path):
    """
     Group the tubes according to the ratio and plot the normalized intensity over time for the different channels at each ratio.
    Insert the image for each ratio into a different sheet in Excel, while generating a separate normalized intensity image for each Tube and inserting it into the corresponding ratio sheet.
    """
    # Grouping using the previous group_tubes_by_specific_ratio function
    ratio_groups = group_tubes_by_specific_ratio(df)

    # Creating Excel Files
    wb = Workbook()

    # Iterate over each ratio
    for ratio, tubes in ratio_groups.items():
        sheet_name = f"Ratio_{ratio[0]}_{ratio[1]}"
        sheet = wb.create_sheet(title=sheet_name)
        plt.figure()  # Creating a new graphic
        plt.title(f"Normalized Intensity Over Time for Ratio {ratio[0]}:{ratio[1]}")

        # Iterate through each test tube and plot a composite graph
        for tube in tubes:
            # Get the data for this tube at ch01 and ch02.
            ch01_data = df[(df["Tube"] == tube) & (df["Channel"] == "ch01")]
            ch02_data = df[(df["Tube"] == tube) & (df["Channel"] == "ch02")]

            # Ensure availability of data
            if not ch01_data.empty and not ch02_data.empty:
                # Sorted data
                ch01_data = ch01_data.sort_values("Time")
                ch02_data = ch02_data.sort_values("Time")

                # Normalization: Setting the maximum intensity to 1
                max_ch01_intensity = ch01_data["Intensity"].max()
                max_ch02_intensity = ch02_data["Intensity"].max()

                ch01_data["Normalized Intensity"] = (
                    ch01_data["Intensity"] / max_ch01_intensity
                )
                ch02_data["Normalized Intensity"] = (
                    ch02_data["Intensity"] / max_ch02_intensity
                )

                # Plotting normalized intensity over time
                plt.plot(
                    ch01_data["Time"],
                    ch01_data["Normalized Intensity"],
                    marker="o",
                    label=f"Tube {tube} - ch01",
                )
                plt.plot(
                    ch02_data["Time"],
                    ch02_data["Normalized Intensity"],
                    marker="x",
                    label=f"Tube {tube} - ch02",
                )

        plt.xlabel("Time")
        plt.ylabel("Normalized Intensity")
        plt.legend()
        plt.grid(True)

        # Save composite image to memory
        img_stream = BytesIO()
        plt.savefig(img_stream, format="png")
        img_stream.seek(0)  # Reset the position of the stream
        img = Image(img_stream)

        # Insert the composite image into the ratio sheet
        sheet.add_image(img, "A1")

        plt.close()  # Close Integrated Graphics

        # Generate a separate image for each tube and insert it into the corresponding ratio sheet
        current_row = 40  # Insert tube image starting in cell A20
        for tube in tubes:
            tube_sheet_title = f"Tube {tube} - Normalized Intensity"
            plt.figure()  # Creating a new graphic
            plt.title(f"Normalized Intensity Over Time for Tube {tube}")

            # Get the tube's data at ch01 and ch02.
            ch01_data = df[(df["Tube"] == tube) & (df["Channel"] == "ch01")]
            ch02_data = df[(df["Tube"] == tube) & (df["Channel"] == "ch02")]

            # Normalization: Setting the maximum intensity to 1
            if not ch01_data.empty and not ch02_data.empty:
                # sort data
                ch01_data = ch01_data.sort_values("Time")
                ch02_data = ch02_data.sort_values("Time")

                max_ch01_intensity = ch01_data["Intensity"].max()
                max_ch02_intensity = ch02_data["Intensity"].max()

                ch01_data["Normalized Intensity"] = (
                    ch01_data["Intensity"] / max_ch01_intensity
                )
                ch02_data["Normalized Intensity"] = (
                    ch02_data["Intensity"] / max_ch02_intensity
                )

                # Plotting normalized intensity over time
                plt.plot(
                    ch01_data["Time"],
                    ch01_data["Normalized Intensity"],
                    marker="o",
                    label=f"ch01",
                )
                plt.plot(
                    ch02_data["Time"],
                    ch02_data["Normalized Intensity"],
                    marker="x",
                    label=f"ch02",
                )

            plt.xlabel("Time")
            plt.ylabel("Normalized Intensity")
            plt.legend()
            plt.grid(True)

            # Save individual tube images to memory
            img_stream_tube = BytesIO()
            plt.savefig(img_stream_tube, format="png")
            img_stream_tube.seek(0)  # Reset the position of the stream
            img_tube = Image(img_stream_tube)

            # Insert tube image into ratio sheet (starting with A20)
            sheet.add_image(img_tube, f"A{current_row}")
            current_row += 40  # Each tube image is spaced 20 lines apart

            plt.close()  # Turning off the tube's graphics
    # Save Excel file
    excel_path = os.path.join(
        excel_path, "fluorescent_analysis_results_intensity_normaliszed_graph.xlsx"
    )
    excel_dir = os.path.dirname(excel_path)

    # If the directory does not exist, create the directory
    if not os.path.exists(excel_dir):
        os.makedirs(excel_dir)
    # Save Excel file
    wb.save(excel_path)
    print(f"Excel file saved to {excel_path}")


root_path = r"D:\thesis\processed 7th\R2\Raw"
# organize_files_by_time_part(root_path)
# df = analysis_intensity_over_time(root_path)
# plot_normalized_intensity_over_time_by_ratio(df,root_path)
# plot_cell_count_over_time_by_ratio(analysis_intensity_over_time(root_path),root_path)
